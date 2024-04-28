import openpyxl
from openpyxl import load_workbook
from datetime import datetime



def read_excel(file_path, sheet_name):

    data = []
    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook(file_path)
        # Select the specified sheet
        sheet = workbook[sheet_name]
        # Iterate over rows in the sheet
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    finally:
        return data


def write_excel(file_path, sheet_name, data):

    try:
        # open existing book
        workbook = openpyxl.Workbook()
        # Select the default active sheet
        sheet = workbook.active
        # Rename the active sheet to sheet name
        sheet.title = sheet_name

        # Write data to the sheet
        for row in data:
            sheet.append(row)
        # Save the workbook to the specified file path
        last_col = sheet.max_column
        new_header = "Test_Result"
        sheet.cell(row=1, column=last_col, value=new_header)

        last_row = sheet.max_row
        current_date = datetime.now().date().strftime("%Y-%m-%d")
        current_time = datetime.now().time().strftime("%H:%M:%S")
        for row_index in range(2, last_row + 1):
            sheet.cell(row=row_index, column=4, value=current_date)
            sheet.cell(row=row_index, column=5, value=current_time)

        workbook.save(file_path)
        print("Data written to Excel file successfully.")
    except Exception as e:
        print(f"Error writing to Excel file: {e}")
